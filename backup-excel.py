#!/usr/bin/env python3
"""
Daily Excel backup - xuất dữ liệu check-in/out + revenue
Chạy cuối ngày (5-6PM), tạo file Excel chi tiết
"""

import sqlite3
import os
from datetime import datetime, timedelta
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

class ExcelBackup:
    def __init__(self, db_path, output_dir="./backups/excel"):
        self.db_path = db_path
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)

        # Excel styling
        self.header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        self.header_font = Font(bold=True, color="FFFFFF")
        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

    def connect_db(self):
        """Kết nối database"""
        return sqlite3.connect(self.db_path)

    def export_checkin_history(self, workbook, days=1):
        """Export lịch sử check-in/out"""
        ws = workbook.create_sheet("Check-in-out")

        conn = self.connect_db()
        cursor = conn.cursor()

        # Query check-in data
        query = """
        SELECT
            DATE(s.date) as Ngày,
            e.fullName as "Tên NV",
            e.employeeId as "Mã NV",
            s.startTime as "Giờ vào ca",
            s.endTime as "Giờ tan ca",
            CASE WHEN s.checkIn THEN TIME(s.checkIn) ELSE '-' END as "Check-in",
            CASE WHEN s.checkOut THEN TIME(s.checkOut) ELSE '-' END as "Check-out",
            CASE
                WHEN s.checkInPhoto IS NOT NULL THEN '✅'
                ELSE '❌'
            END as "Ảnh In",
            CASE
                WHEN s.checkOutPhoto IS NOT NULL THEN '✅'
                ELSE '❌'
            END as "Ảnh Out",
            s.branch as "Chi nhánh"
        FROM shifts s
        LEFT JOIN employees e ON s.employeeId = e.id
        WHERE DATE(s.date) >= DATE('now', '-' || ? || ' days')
        ORDER BY s.date DESC, s.employeeId
        """

        cursor.execute(query, (days,))
        data = cursor.fetchall()

        # Headers
        headers = [description[0] for description in cursor.description]
        ws.append(headers)

        # Style headers
        for cell in ws[1]:
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = self.border

        # Add data
        for row in data:
            ws.append(row)
            for cell in ws[ws.max_row]:
                cell.border = self.border
                cell.alignment = Alignment(horizontal='center')

        # Auto-width columns
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            ws.column_dimensions[column_letter].width = min(max_length + 2, 30)

        conn.close()
        print(f"✅ Exported {len(data)} check-in records")

    def export_daily_summary(self, workbook, date=None):
        """Tóm tắt ngày hôm nay"""
        if date is None:
            date = datetime.now().strftime('%Y-%m-%d')

        ws = workbook.create_sheet("Tóm tắt hôm nay", 0)

        conn = self.connect_db()
        cursor = conn.cursor()

        # Summary statistics
        cursor.execute("""
        SELECT
            COUNT(DISTINCT employeeId) as "Số NV check-in",
            COUNT(*) as "Tổng shifts",
            SUM(CASE WHEN checkInPhoto IS NOT NULL THEN 1 ELSE 0 END) as "Ảnh check-in",
            SUM(CASE WHEN checkOutPhoto IS NOT NULL THEN 1 ELSE 0 END) as "Ảnh check-out"
        FROM shifts
        WHERE date = ?
        """, (date,))

        stats = cursor.fetchone()

        # Write summary
        ws['A1'] = f"📊 Tóm tắt {date}"
        ws['A1'].font = Font(bold=True, size=14)

        row = 3
        ws[f'A{row}'] = "Số nhân viên check-in:"
        ws[f'B{row}'] = stats[0] or 0

        row += 1
        ws[f'A{row}'] = "Tổng shifts:"
        ws[f'B{row}'] = stats[1] or 0

        row += 1
        ws[f'A{row}'] = "Ảnh check-in:"
        ws[f'B{row}'] = stats[2] or 0

        row += 1
        ws[f'A{row}'] = "Ảnh check-out:"
        ws[f'B{row}'] = stats[3] or 0

        # Detailed by employee
        row += 3
        ws[f'A{row}'] = "Chi tiết theo nhân viên"
        ws[f'A{row}'].font = Font(bold=True, size=12)

        row += 1
        headers = ["Tên NV", "Mã NV", "Check-in", "Check-out", "Ảnh In", "Ảnh Out"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col)
            cell.value = header
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.border = self.border

        # Query employee check-in
        cursor.execute("""
        SELECT
            e.fullName,
            e.employeeId,
            CASE WHEN s.checkIn THEN '✅' ELSE '❌' END,
            CASE WHEN s.checkOut THEN '✅' ELSE '❌' END,
            CASE WHEN s.checkInPhoto THEN '✅' ELSE '❌' END,
            CASE WHEN s.checkOutPhoto THEN '✅' ELSE '❌' END
        FROM shifts s
        LEFT JOIN employees e ON s.employeeId = e.id
        WHERE s.date = ?
        ORDER BY e.fullName
        """, (date,))

        for record in cursor.fetchall():
            row += 1
            for col, value in enumerate(record, 1):
                cell = ws.cell(row=row, column=col)
                cell.value = value
                cell.border = self.border
                cell.alignment = Alignment(horizontal='center')

        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 12

        conn.close()
        print(f"✅ Summary for {date} exported")

    def export_employee_stats(self, workbook, days=7):
        """Thống kê nhân viên 7 ngày gần đây"""
        ws = workbook.create_sheet("Thống kê nhân viên")

        conn = self.connect_db()
        cursor = conn.cursor()

        cursor.execute("""
        SELECT
            e.fullName,
            e.employeeId,
            COUNT(*) as "Số ca",
            SUM(CASE WHEN s.checkIn IS NOT NULL THEN 1 ELSE 0 END) as "Check-in",
            SUM(CASE WHEN s.checkOut IS NOT NULL THEN 1 ELSE 0 END) as "Check-out",
            SUM(CASE WHEN s.checkInPhoto IS NOT NULL THEN 1 ELSE 0 END) as "Ảnh In",
            SUM(CASE WHEN s.checkOutPhoto IS NOT NULL THEN 1 ELSE 0 END) as "Ảnh Out"
        FROM shifts s
        LEFT JOIN employees e ON s.employeeId = e.id
        WHERE s.date >= DATE('now', '-' || ? || ' days')
        GROUP BY e.id, e.fullName, e.employeeId
        ORDER BY COUNT(*) DESC
        """, (days,))

        headers = ["Tên NV", "Mã NV", "Số ca", "Check-in", "Check-out", "Ảnh In", "Ảnh Out"]
        ws.append(headers)

        for cell in ws[1]:
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.alignment = Alignment(horizontal='center')
            cell.border = self.border

        for row in cursor.fetchall():
            ws.append(row)
            for cell in ws[ws.max_row]:
                cell.border = self.border
                if ws.cell(ws.max_row, ws.max_column).column > 2:
                    cell.alignment = Alignment(horizontal='center')

        # Auto-width
        for i, column in enumerate(ws.columns, 1):
            ws.column_dimensions[get_column_letter(i)].width = 15

        conn.close()
        print(f"✅ Employee stats for {days} days exported")

    def export_all(self, days=1):
        """Export tất cả thành 1 file Excel"""
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = self.output_dir / f"backup_{timestamp}.xlsx"

        workbook = openpyxl.Workbook()
        workbook.remove(workbook.active)  # Remove default sheet

        print(f"\n📝 Creating backup: {filename}")

        # Export các sheet
        self.export_daily_summary(workbook)
        self.export_checkin_history(workbook, days=days)
        self.export_employee_stats(workbook, days=7)

        # Save
        workbook.save(filename)
        print(f"✅ Excel backup saved: {filename}")
        print(f"📊 File size: {filename.stat().st_size / 1024:.1f} KB")

        return filename

if __name__ == '__main__':
    # Configuration
    DB_PATH = '/Users/taidoan/Desktop/sinhtoooo/sinhto/data/database.sqlite'

    # Create backup
    backup = ExcelBackup(DB_PATH)
    file = backup.export_all(days=1)  # Last 1 day

    print(f"\n✅ Backup completed!")
    print(f"   Location: {file}")
